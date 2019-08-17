#!/usr/bin/env python
# -*- coding:utf-8 -*-


import sys
import pymongo
import time
import execjs
import requests
import re
from openpyxl import load_workbook
from openpyxl import Workbook
from pyquery import PyQuery as pq
from PyQt5.QtWidgets import QApplication, QPushButton, QLabel, QFileDialog, QInputDialog, QTextBrowser, QFrame
from PyQt5.QtWidgets import QMessageBox, QLineEdit, QDialog, QProgressBar
from PyQt5.QtGui import QIcon, QPixmap, QPalette, QBrush, QMovie
from PyQt5.QtCore import Qt, QThread, pyqtSignal
# import PyQt5.sip


time_start = None
time_stop = None


LOGIN_PWDS = ['201707030103']
login_url = 'http://202.115.133.173:805/Common/Handler/UserLogin.ashx'
score_url = 'http://202.115.133.173:805/SearchInfo/Score/ScoreList.aspx'
person = {}


jsstr = '''
function UserLogin(userName, pwd) {

    var sign = new Date().getTime();
    var user = userName.trim();
    var signedpwd = hex_md5(user + sign + hex_md5(pwd.trim()));

    var data = {Action: "Login", userName: user, pwd: signedpwd, sign: sign};

    return data;
}

var hexcase = 0;  /* hex output format. 0 - lowercase; 1 - uppercase        */
var b64pad  = ""; /* base-64 pad character. "=" for strict RFC compliance   */
var chrsz   = 8;  /* bits per input character. 8 - ASCII; 16 - Unicode      */

function hex_md5(s){ return binl2hex(core_md5(str2binl(s), s.length * chrsz));}
function b64_md5(s){ return binl2b64(core_md5(str2binl(s), s.length * chrsz));}
function str_md5(s){ return binl2str(core_md5(str2binl(s), s.length * chrsz));}
function hex_hmac_md5(key, data) { return binl2hex(core_hmac_md5(key, data)); }
function b64_hmac_md5(key, data) { return binl2b64(core_hmac_md5(key, data)); }
function str_hmac_md5(key, data) { return binl2str(core_hmac_md5(key, data)); }

function md5_vm_test()
{
  return hex_md5("abc") == "900150983cd24fb0d6963f7d28e17f72";
}

function core_md5(x, len)
{
  x[len >> 5] |= 0x80 << ((len) % 32);
  x[(((len + 64) >>> 9) << 4) + 14] = len;

  var a =  1732584193;
  var b = -271733879;
  var c = -1732584194;
  var d =  271733878;

  for(var i = 0; i < x.length; i += 16)
  {
    var olda = a;
    var oldb = b;
    var oldc = c;
    var oldd = d;

    a = md5_ff(a, b, c, d, x[i+ 0], 7 , -680876936);
    d = md5_ff(d, a, b, c, x[i+ 1], 12, -389564586);
    c = md5_ff(c, d, a, b, x[i+ 2], 17,  606105819);
    b = md5_ff(b, c, d, a, x[i+ 3], 22, -1044525330);
    a = md5_ff(a, b, c, d, x[i+ 4], 7 , -176418897);
    d = md5_ff(d, a, b, c, x[i+ 5], 12,  1200080426);
    c = md5_ff(c, d, a, b, x[i+ 6], 17, -1473231341);
    b = md5_ff(b, c, d, a, x[i+ 7], 22, -45705983);
    a = md5_ff(a, b, c, d, x[i+ 8], 7 ,  1770035416);
    d = md5_ff(d, a, b, c, x[i+ 9], 12, -1958414417);
    c = md5_ff(c, d, a, b, x[i+10], 17, -42063);
    b = md5_ff(b, c, d, a, x[i+11], 22, -1990404162);
    a = md5_ff(a, b, c, d, x[i+12], 7 ,  1804603682);
    d = md5_ff(d, a, b, c, x[i+13], 12, -40341101);
    c = md5_ff(c, d, a, b, x[i+14], 17, -1502002290);
    b = md5_ff(b, c, d, a, x[i+15], 22,  1236535329);

    a = md5_gg(a, b, c, d, x[i+ 1], 5 , -165796510);
    d = md5_gg(d, a, b, c, x[i+ 6], 9 , -1069501632);
    c = md5_gg(c, d, a, b, x[i+11], 14,  643717713);
    b = md5_gg(b, c, d, a, x[i+ 0], 20, -373897302);
    a = md5_gg(a, b, c, d, x[i+ 5], 5 , -701558691);
    d = md5_gg(d, a, b, c, x[i+10], 9 ,  38016083);
    c = md5_gg(c, d, a, b, x[i+15], 14, -660478335);
    b = md5_gg(b, c, d, a, x[i+ 4], 20, -405537848);
    a = md5_gg(a, b, c, d, x[i+ 9], 5 ,  568446438);
    d = md5_gg(d, a, b, c, x[i+14], 9 , -1019803690);
    c = md5_gg(c, d, a, b, x[i+ 3], 14, -187363961);
    b = md5_gg(b, c, d, a, x[i+ 8], 20,  1163531501);
    a = md5_gg(a, b, c, d, x[i+13], 5 , -1444681467);
    d = md5_gg(d, a, b, c, x[i+ 2], 9 , -51403784);
    c = md5_gg(c, d, a, b, x[i+ 7], 14,  1735328473);
    b = md5_gg(b, c, d, a, x[i+12], 20, -1926607734);

    a = md5_hh(a, b, c, d, x[i+ 5], 4 , -378558);
    d = md5_hh(d, a, b, c, x[i+ 8], 11, -2022574463);
    c = md5_hh(c, d, a, b, x[i+11], 16,  1839030562);
    b = md5_hh(b, c, d, a, x[i+14], 23, -35309556);
    a = md5_hh(a, b, c, d, x[i+ 1], 4 , -1530992060);
    d = md5_hh(d, a, b, c, x[i+ 4], 11,  1272893353);
    c = md5_hh(c, d, a, b, x[i+ 7], 16, -155497632);
    b = md5_hh(b, c, d, a, x[i+10], 23, -1094730640);
    a = md5_hh(a, b, c, d, x[i+13], 4 ,  681279174);
    d = md5_hh(d, a, b, c, x[i+ 0], 11, -358537222);
    c = md5_hh(c, d, a, b, x[i+ 3], 16, -722521979);
    b = md5_hh(b, c, d, a, x[i+ 6], 23,  76029189);
    a = md5_hh(a, b, c, d, x[i+ 9], 4 , -640364487);
    d = md5_hh(d, a, b, c, x[i+12], 11, -421815835);
    c = md5_hh(c, d, a, b, x[i+15], 16,  530742520);
    b = md5_hh(b, c, d, a, x[i+ 2], 23, -995338651);

    a = md5_ii(a, b, c, d, x[i+ 0], 6 , -198630844);
    d = md5_ii(d, a, b, c, x[i+ 7], 10,  1126891415);
    c = md5_ii(c, d, a, b, x[i+14], 15, -1416354905);
    b = md5_ii(b, c, d, a, x[i+ 5], 21, -57434055);
    a = md5_ii(a, b, c, d, x[i+12], 6 ,  1700485571);
    d = md5_ii(d, a, b, c, x[i+ 3], 10, -1894986606);
    c = md5_ii(c, d, a, b, x[i+10], 15, -1051523);
    b = md5_ii(b, c, d, a, x[i+ 1], 21, -2054922799);
    a = md5_ii(a, b, c, d, x[i+ 8], 6 ,  1873313359);
    d = md5_ii(d, a, b, c, x[i+15], 10, -30611744);
    c = md5_ii(c, d, a, b, x[i+ 6], 15, -1560198380);
    b = md5_ii(b, c, d, a, x[i+13], 21,  1309151649);
    a = md5_ii(a, b, c, d, x[i+ 4], 6 , -145523070);
    d = md5_ii(d, a, b, c, x[i+11], 10, -1120210379);
    c = md5_ii(c, d, a, b, x[i+ 2], 15,  718787259);
    b = md5_ii(b, c, d, a, x[i+ 9], 21, -343485551);

    a = safe_add(a, olda);
    b = safe_add(b, oldb);
    c = safe_add(c, oldc);
    d = safe_add(d, oldd);
  }
  return Array(a, b, c, d);

}

function md5_cmn(q, a, b, x, s, t)
{
  return safe_add(bit_rol(safe_add(safe_add(a, q), safe_add(x, t)), s),b);
}
function md5_ff(a, b, c, d, x, s, t)
{
  return md5_cmn((b & c) | ((~b) & d), a, b, x, s, t);
}
function md5_gg(a, b, c, d, x, s, t)
{
  return md5_cmn((b & d) | (c & (~d)), a, b, x, s, t);
}
function md5_hh(a, b, c, d, x, s, t)
{
  return md5_cmn(b ^ c ^ d, a, b, x, s, t);
}
function md5_ii(a, b, c, d, x, s, t)
{
  return md5_cmn(c ^ (b | (~d)), a, b, x, s, t);
}

function core_hmac_md5(key, data)
{
  var bkey = str2binl(key);
  if(bkey.length > 16) bkey = core_md5(bkey, key.length * chrsz);

  var ipad = Array(16), opad = Array(16);
  for(var i = 0; i < 16; i++)
  {
    ipad[i] = bkey[i] ^ 0x36363636;
    opad[i] = bkey[i] ^ 0x5C5C5C5C;
  }

  var hash = core_md5(ipad.concat(str2binl(data)), 512 + data.length * chrsz);
  return core_md5(opad.concat(hash), 512 + 128);
}

function safe_add(x, y)
{
  var lsw = (x & 0xFFFF) + (y & 0xFFFF);
  var msw = (x >> 16) + (y >> 16) + (lsw >> 16);
  return (msw << 16) | (lsw & 0xFFFF);
}

function bit_rol(num, cnt)
{
  return (num << cnt) | (num >>> (32 - cnt));
}

function str2binl(str)
{
  var bin = Array();
  var mask = (1 << chrsz) - 1;
  for(var i = 0; i < str.length * chrsz; i += chrsz)
    bin[i>>5] |= (str.charCodeAt(i / chrsz) & mask) << (i%32);
  return bin;
}

function binl2str(bin)
{
  var str = "";
  var mask = (1 << chrsz) - 1;
  for(var i = 0; i < bin.length * 32; i += chrsz)
    str += String.fromCharCode((bin[i>>5] >>> (i % 32)) & mask);
  return str;
}


function binl2hex(binarray)
{
  var hex_tab = hexcase ? "0123456789ABCDEF" : "0123456789abcdef";
  var str = "";
  for(var i = 0; i < binarray.length * 4; i++)
  {
    str += hex_tab.charAt((binarray[i>>2] >> ((i%4)*8+4)) & 0xF) +
           hex_tab.charAt((binarray[i>>2] >> ((i%4)*8  )) & 0xF);
  }
  return str;
}

function binl2b64(binarray)
{
  var tab = "ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789+/";
  var str = "";
  for(var i = 0; i < binarray.length * 4; i += 3)
  {
    var triplet = (((binarray[i   >> 2] >> 8 * ( i   %4)) & 0xFF) << 16)
                | (((binarray[i+1 >> 2] >> 8 * ((i+1)%4)) & 0xFF) << 8 )
                |  ((binarray[i+2 >> 2] >> 8 * ((i+2)%4)) & 0xFF);
    for(var j = 0; j < 4; j++)
    {
      if(i * 8 + j * 6 > binarray.length * 32) str += b64pad;
      else str += tab.charAt((triplet >> 6*(3-j)) & 0x3F);
    }
  }
  return str;
}

'''


n = None
accounts = None
passwords = None


class LoginDialog(QDialog):
    login_signal = pyqtSignal(str)

    def __init__(self):
        super(LoginDialog, self).__init__()
        self.initUI()

    def initUI(self):
        self.setFixedSize(350, 250)
        self.setWindowIcon(QIcon('app.ico'))
        self.setWindowTitle('登录')
        self.setWindowFlags(Qt.FramelessWindowHint)

        self.picLabel = QLabel('', self)
        self.picLabel.setFrameStyle(QFrame.Panel | QFrame.Sunken)
        self.picLabel.setGeometry(0, 0, 350, 140)
        self.movie = QMovie('杜甫.gif')
        if self.movie.isValid():
            self.picLabel.setMovie(self.movie)
            self.picLabel.setScaledContents(True)
            self.movie.start()
        else:
            self.picLabel.setPixmap(QPixmap('.\CDUT_pic_1.jpeg'))

        self.pwdEdit = QLineEdit(self)
        self.pwdEdit.setGeometry(80, 160, 200, 30)
        self.pwdEdit.setPlaceholderText('请输入密码')
        self.pwdEdit.setEchoMode(QLineEdit.Password)

        self.submitBtn = QPushButton('登录', self)
        self.submitBtn.setGeometry(82, 210, 60, 30)
        self.submitBtn.setStyleSheet("QPushButton{color:black}"
                                      "QPushButton:hover{color:red}"
                                      "QPushButton{background-color:lightgreen}"
                                      "QPushButton{border:2px}"
                                      "QPushButton{border-radius:15px}"
                                      "QPushButton{padding:2px 4px}")

        self.closeBtn = QPushButton('关闭', self)
        self.closeBtn.setGeometry(220, 210, 60, 30)
        self.closeBtn.setStyleSheet("QPushButton{color:black}"
                                     "QPushButton:hover{color:red}"
                                     "QPushButton{background-color:lightgreen}"
                                     "QPushButton{border:2px}"
                                     "QPushButton{border-radius:15px}"
                                     "QPushButton{padding:2px 4px}")

        self.submitBtn.clicked.connect(self.submitLogin)
        self.closeBtn.clicked.connect(self.closeLogin)

        self.show()

    def submitLogin(self):
        self.text = self.pwdEdit.text()

        if self.text not in LOGIN_PWDS:
            msgBox = QMessageBox()
            msgBox.setWindowOpacity(0.8)
            msgBox.setWindowTitle('错误')
            msgBox.setWindowIcon(QIcon('app.ico'))
            msgBox.setIcon(QMessageBox.Critical)
            msgBox.setText('密码输入错误！')
            msgBox.setInformativeText('请联系QQ:<span style="color: red">1792575431</span>')
            msgBox.addButton('确定', QMessageBox.AcceptRole)
            msgBox.exec_()
            self.pwdEdit.clear()
            self.login_signal.emit('fail')

        else:
            self.login_signal.emit('success')
            self.close()

    def closeLogin(self):
        self.close()

    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton:
            self.m_flag = True
            self.m_Position = event.globalPos() - self.pos()
            event.accept()

    def mouseMoveEvent(self, QMouseEvent):
        if Qt.LeftButton and self.m_flag:
            self.move(QMouseEvent.globalPos() - self.m_Position)
            QMouseEvent.accept()

    def mouseReleaseEvent(self, QMouseEvent):
        self.m_flag = False


class GetScore(QDialog):
    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        self.setFixedSize(350, 600)
        self.setWindowFlags(Qt.WindowMinimizeButtonHint | Qt.WindowCloseButtonHint)
        self.setWindowTitle('成理成绩查询')
        self.setWindowIcon(QIcon('.\\app.ico'))
        window_pale = QPalette()
        window_pale.setBrush(self.backgroundRole(), QBrush(QPixmap('.\pic3.jpg')))
        self.setPalette(window_pale)
        self.setWindowOpacity(1)
        self.setStyleSheet("QPushButton{background: transparent} QTextBrowser{background: transparent; border: 0}")

        self.aboutBtn = QPushButton('关于', self)
        self.aboutBtn.setGeometry(145, 580, 40, 25)
        self.aboutBtn.setStyleSheet('color: red; text-decoration: underline')

        self.fileLabel = QLabel('信息文件：', self)
        self.fileLabel.move(20, 20)
        self.fileValueLabel = QLabel('账号第1列，密码第2列', self)
        self.fileValueLabel.setWordWrap(True)
        self.fileValueLabel.setStyleSheet("color: blue")
        self.fileValueLabel.setGeometry(100, 15, 160, 25)
        self.fileBtn = QPushButton('...', self)
        self.fileBtn.setStyleSheet("QPushButton:hover{color: orange}")
        self.fileBtn.setGeometry(300, 15, 40, 30)

        self.termLabel = QLabel('学期：', self)
        self.termLabel.move(20, 80)
        self.termValueLabel = QLabel('', self)
        self.termValueLabel.setStyleSheet("color: blue")
        self.termValueLabel.setGeometry(100, 75, 160, 25)
        self.termBtn = QPushButton('...', self)
        self.termBtn.setStyleSheet("QPushButton:hover{color: orange}")
        self.termBtn.setGeometry(300, 75, 40, 30)

        self.saveLabel = QLabel('存储位置：', self)
        self.saveLabel.move(20, 140)
        self.saveValueLabel = QLabel('', self)
        self.saveValueLabel.setStyleSheet("color: blue")
        self.saveValueLabel.setGeometry(100, 135, 160, 25)
        self.saveBtn = QPushButton('...', self)
        self.saveBtn.setStyleSheet("QPushButton:hover{color: orange}")
        self.saveBtn.setGeometry(300, 135, 40, 30)

        self.infoLabel = QLabel('执行情况↓↓↓', self)
        self.infoLabel.move(20, 200)
        self.infoText = QTextBrowser(self)
        self.infoText.setContextMenuPolicy(Qt.NoContextMenu)
        self.infoText.move(20, 240)

        self.executeBtn = QPushButton('执行', self)
        self.executeBtn.setStyleSheet("QPushButton:hover{color: green}"
                                      "")
        self.executeBtn.setGeometry(145, 480, 40, 25)

        self.progress = QProgressBar(self)
        self.progress.setMinimum(0)
        self.progress.setMaximum(0)
        self.progress.setGeometry(65, 520, 200, 20)
        self.progress.setStyleSheet(
            "QProgressBar{border-radius: 10px; border-radius: 10px; text-align: center; color: red}"
            "QProgressBar:chunk{border: 1px solid grey; border-radius: 5px; background-color: grey;}")

        self.fileBtn.clicked.connect(self.fileopen)
        self.termBtn.clicked.connect(self.chooseterm)
        self.saveBtn.clicked.connect(self.filestore)
        self.aboutBtn.clicked.connect(self.about)
        self.executeBtn.clicked.connect(self.thread_start)

    def showOrcloseDialog(self, info):
        if info == 'success':
            if not self.isVisible():
                self.show()
        elif info == 'fail':
            if self.isVisible():
                self.close()

    def changetxt(self, file_inf):
        if file_inf == 'stop':
            self.executeBtn.setEnabled(True)
            self.infoText.append('\n耗时：%.2fs' % (time_stop - time_start))
        elif file_inf == '<span style="color: red">请选择正确的学期！</span>':
            self.infoText.append(file_inf)
            self.executeBtn.setEnabled(True)
        else:
            self.infoText.append(file_inf)
            value = self.progress.value()
            value += 1
            self.progress.setValue(value)

    def showerror(self, info):
        if info == 0:
            self.infoText.append('打开文件失败！\n')
        else:
            self.infoText.append('打开文件成功！\n')

    def thread_start(self):
        global time_start
        global person
        time_start = time.time()
        person = {}

        if self.fileValueLabel.text() == '':
            self.infoText.append('<span style="color: red">请选择信息文件！</span>')
            return
        elif self.termValueLabel.text() == '':
            self.infoText.append('<span style="color :red">请选择学期！</span>')
            return
        elif self.saveValueLabel.text() == '':
            self.infoText.append('<span style="color: red">请选择存储位置！</span>')
            return

        global n
        self.progress.setMaximum(n)
        self.progress.setValue(0)

        self.executeBtn.setEnabled(False)
        self.term = self.termValueLabel.text()
        self.infoText.clear()
        self.thread_1 = Thread_get(term=self.term, file=self.fname[0])
        self.thread_1.file_changed_signal.connect(self.changetxt)
        self.thread_1.start()

    def fileopen(self):
        fname = QFileDialog.getOpenFileName(self, '打开文件', './', filter='XLSX files (*.xlsx)' or 'XLS files (*.xls)')
        if fname[0]:
            filename = []
            i = -1
            while True:
                if fname[0][i] == '/':
                    break
                filename.append(fname[0][i])
                i -= 1
            filename = ''.join(filename)
            self.fileValueLabel.setText(filename[::-1])

            try:
                thread_open = Thread_open(file=fname[0])
                thread_open.info.connect(self.showerror)
                thread_open.start()
            except Exception:
                self.infoText.append('<span style="color: red">打开文件异常！</span>')

    def chooseterm(self):
        inputdialog = QInputDialog(self)
        items = ['大一上期', '大一下期', '大二上期', '大二下期', '大三上期', '大三下期', '大四上期', '大四下期']
        inputdialog.setComboBoxItems(items)
        inputdialog.setWindowTitle('学期')
        inputdialog.setLabelText('选择学期：')
        inputdialog.setOkButtonText('确定')
        inputdialog.setCancelButtonText('取消')
        inputdialog.setWindowOpacity(0.7)
        ok = inputdialog.exec_()
        if ok:
            self.termValueLabel.setText(str(inputdialog.textValue()))

    def filestore(self):
        self.fname = QFileDialog.getSaveFileName(self, '保存文件', directory='./', filter='XLSX files (*.xlsx)')
        if self.fname[0]:
            filename = []
            i = -1
            while True:
                if self.fname[0][i] == '/':
                    break
                filename.append(self.fname[0][i])
                i -= 1
            filename = ''.join(filename)
            self.saveValueLabel.setText(filename[::-1])

    def about(self):
        msgBox = QMessageBox(QMessageBox.NoIcon, '关于', '<b>钦哥出品</b>')
        msgBox.setWindowOpacity(0.7)
        msgBox.addButton('确定', QMessageBox.AcceptRole)
        msgBox.setWindowIcon(QIcon('.\\app.ico'))
        msgBox.setInformativeText('如有疑问请联系QQ:<span style="color: red">1792575431</span>')
        msgBox.setIconPixmap(QPixmap('.\My_logo.jpg'))
        msgBox.exec_()


class Thread_open(QThread):
    info = pyqtSignal(int)

    def __init__(self, file=None, parent=None):
        self.file = file
        super(Thread_open, self).__init__(parent)

    def __del__(self):
        self.wait()

    def run(self):
        try:
            global n
            global accounts
            global passwords
            wb = load_workbook(self.file)
            sheet_names = wb.sheetnames
            sheet = wb[sheet_names[0]]
            accounts = []
            passwords = []
            for cell in list(sheet.columns)[0]:
                accounts.append(cell.value)
            for cell in list(sheet.columns)[1]:
                passwords.append(cell.value)
            n = len(accounts)
            self.info.emit(1)
        except Exception:
            self.info.emit(0)


class Thread_get(QThread):
    file_changed_signal = pyqtSignal(str)

    def __init__(self, term=None, file=None, parent=None):
        super(Thread_get, self).__init__(parent)
        global n
        global accounts
        global passwords
        self.n = n
        self.accounts = accounts
        self.passwords = passwords
        self.file = file
        self.errorCnt = 0
        items = {'大一上期': [0, 1], '大一下期': [0, 2], '大二上期': [1, 1], '大二下期': [1, 2], '大三上期': [2, 1], '大三下期': [2, 2],
        '大四上期': [3, 1], '大四下期': [3, 2]}
        for item in items.items():
            if term == item[0]:
                self.term = str(int(self.accounts[0][:4]) + item[1][0]) + '0' + str(item[1][1])
                break

    def __del__(self):
        self.wait()

    def run(self):
        global person

        zh = {'优': '95', '良': '85', '中': '75', '及格': '65', '不及格': '55'}
        ctx = execjs.compile(jsstr)
        params = ctx.call('UserLogin', self.accounts[0], self.passwords[0])
        session = requests.Session()
        session.post(login_url, params)
        s = session.get(score_url)
        html = s.text
        doc = pq(html)
        courses = doc('.score_right_infor_list.listUl')
        courses = courses.children()
        total_terms = []
        for item in courses.items():
            term = item.children('.floatDiv20').text().strip()
            total_terms.append(term)
        if self.term not in total_terms:
            self.file_changed_signal.emit('<span style="color: red">请选择正确的学期！</span>')
            return

        for i in range(self.n):
            try:
                name = ''
                score = {}
                credit = {}
                isRX = {}
                allScore = {}
                allCredit = {}
                allIsRX = {}
                score2 = {}
                credit2 = {}
                isRX2 = {}
                loop = 0
                while len(score) == 0:
                    loop += 1
                    ctx = execjs.compile(jsstr)
                    params = ctx.call('UserLogin', self.accounts[i], self.passwords[i])
                    session = requests.Session()
                    session.post(login_url, params)
                    s = session.get(score_url)
                    html = s.text
                    doc = pq(html)
                    nameTxt = doc('.ico_user.ico_blue').text()
                    name = re.search('(.*?)\s', nameTxt, re.S)
                    courses = doc('.score_right_infor_list.listUl')
                    courses = courses.children()
                    for item in courses.items():
                        term = item.children('.floatDiv20').text().strip()
                        if term != '学期':
                            title = item.find('div:nth-child(3)').text().strip()
                            cj = item.find('div:nth-child(6)').text().strip()
                            xf = item.find('div:nth-child(5)').text().strip()
                            bh = item.find('div:nth-child(2)').text().strip()
                            if bh[0] == 'R' and bh[1] == 'X':
                                bh_ = 1
                            else:
                                bh_ = 0

                            if cj in zh.keys():
                                cj = zh[cj]
                            allScore[title] = float(cj)
                            allCredit[title] = float(xf)
                            allIsRX[title] = bh_
                        if term[0:4] == self.term[0:4]:
                            title = item.find('div:nth-child(3)').text().strip()
                            cj = item.find('div:nth-child(6)').text().strip()
                            xf = item.find('div:nth-child(5)').text().strip()
                            bh = item.find('div:nth-child(2)').text().strip()
                            if bh[0] == 'R' and bh[1] == 'X':
                                bh_ = 1
                            else:
                                bh_ = 0

                            if cj in zh.keys():
                                cj = zh[cj]
                            score2[title] = float(cj)
                            credit2[title] = float(xf)
                            isRX2[title] = bh_
                        if term == self.term:
                            title = item.find('div:nth-child(3)').text().strip()
                            cj = item.find('div:nth-child(6)').text().strip()
                            xf = item.find('div:nth-child(5)').text().strip()
                            bh = item.find('div:nth-child(2)').text().strip()
                            if bh[0] == 'R' and bh[1] == 'X':
                                bh_ = 1
                            else:
                                bh_ = 0

                            if cj in zh.keys():
                                cj = zh[cj]
                            score[title] = float(cj)
                            credit[title] = float(xf)
                            isRX[title] = bh_

                    allXf = 0
                    qh = 0
                    allXf2 = 0
                    qh2 = 0
                    allXf3 = 0
                    qh3 = 0
                    for item in score.keys():
                        if isRX[item] == 0:
                            allXf += credit[item]
                            qh += (score[item] / 10 - 5) * credit[item]

                    jd = qh / allXf
                    score['aaa本学期绩点(不含选修)'] = float(jd)

                    for item in score2.keys():
                        if isRX2[item] == 0:
                            allXf3 += credit2[item]
                            qh3 += (score2[item] / 10 - 5) * credit2[item]

                    jd = qh3 / allXf3
                    score['aab本学年绩点(不含选修)'] = float(jd)

                    for item in allScore.keys():
                        if allIsRX[item] == 0:
                            allXf2 += allCredit[item]
                            qh2 += (allScore[item] / 10 - 5) * allCredit[item]

                    jd = qh2 / allXf2
                    score['aac总绩点(不含选修)'] = float(jd)

                    while loop > 20:
                        raise Exception
                person[name[0].strip()] = score

                self.file_changed_signal.emit('{} 存储成功！'.format(self.accounts[i]))

            except Exception:
                self.errorCnt += 1
                self.file_changed_signal.emit('{} 存储<span style="color: red">失败</span>！'.format(self.accounts[i]))

        self.save_to_excel = Thread_save_to_excel(self.file)
        self.save_to_excel.start()

        self.file_changed_signal.emit('\n\n执行完毕！')
        self.file_changed_signal.emit('失败：<span style="color: red">{}</span> 处'.format(self.errorCnt))
        global time_stop
        time_stop = time.time()
        self.file_changed_signal.emit('stop')


class Thread_save_to_excel(QThread):
    def __init__(self, file=None, parent=None):
        super(Thread_save_to_excel, self).__init__(parent)
        global person
        self.person = person
        self.file = file

    def __del__(self):
        self.wait()

    def run(self):
        max_courses = []
        for course in self.person.values():
            max_courses.extend(list(course.keys()))
        max_courses = list(set(max_courses))
        max_courses.sort()
        wb = Workbook()
        ws = wb.active
        ws.append(['姓名'] + max_courses)
        for name, score in person.items():
            name = [name]
            for c in max_courses:
                if c not in score.keys():
                    score[c] = ''
            score = dict(sorted(score.items(), key=lambda x: x[0]))
            ws.append(name + list(score.values()))
        ws['B1'] = '本学期绩点(不含选修)'
        ws['C1'] = '本学年绩点(不含选修)'
        ws['D1'] = '总绩点(不含选修)'
        wb.save(self.file)


MONGO_URI = 'localhost'
MONGO_DB = 'My_item_1'
MONGO_COLLECTION = 'score_1'


class Thread_save_to_mongodb(QThread):
    def __init__(self, score=None, file=None, parent=None):
        super(Thread_save_to_mongodb, self).__init__(parent)
        self.score = score
        self.file = file
        self.client = pymongo.MongoClient(MONGO_URI)
        self.db = self.client[MONGO_DB]
        self.collection = MONGO_COLLECTION

    def __del__(self):
        self.wait()

    def run(self):
        self.db[self.collection].insert(dict(self.score))


if __name__ == '__main__':
    app = QApplication(sys.argv)
    lg = LoginDialog()
    gs = GetScore()
    lg.login_signal.connect(gs.showOrcloseDialog)
    sys.exit(app.exec_())
